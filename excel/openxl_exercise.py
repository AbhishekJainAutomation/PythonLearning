import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sh = wb['Sheet1']
    # cell = sh['a1']
    # print(sh.max_row)

    for row in range(2, sh.max_row+1 ):
        cell = sh.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sh.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sh,
                       min_row=2,
                       max_row=4,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sh.add_chart(chart, 'e2')
    wb.save(filename)